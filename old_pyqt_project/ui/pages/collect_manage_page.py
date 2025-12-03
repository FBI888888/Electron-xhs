"""
采集管理页面
"""
import re
import json
import time
from datetime import datetime
from pathlib import Path
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QHeaderView, QFileDialog, QDialog,
                             QTextEdit, QLabel)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from qfluentwidgets import (PushButton, PrimaryPushButton, InfoBar, InfoBarPosition,
                           FluentIcon as FIF, MessageBox)

from api.blogger_api import BloggerAPI


class CollectThread(QThread):
    """采集线程"""
    # 信号：index, success, message, data
    item_collected = pyqtSignal(int, bool, str, dict)
    # 信号：index, status
    status_changed = pyqtSignal(int, str)
    # 所有采集完成信号
    all_finished = pyqtSignal()
    # 账号使用信号：account_index, used_count
    account_used = pyqtSignal(int, int)
    
    def __init__(self, collect_items, accounts, max_count=9999, selected_fields=None):
        super().__init__()
        self.collect_items = collect_items
        self.accounts = accounts  # 账号列表
        self.max_count = max_count  # 每个账号的最大使用次数
        self.selected_fields = selected_fields or []  # 用户选择的数据表现字段
        self.is_running = True
        self.is_paused = False
        self.current_account_index = 0  # 当前使用的账号索引
        
    def get_next_available_account(self):
        """获取下一个可用账号"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        # 尝试找到一个未超过最大使用次数的账号
        attempts = 0
        while attempts < len(self.accounts):
            account = self.accounts[self.current_account_index]
            
            # 检查是否为今天的使用记录
            if account.get('last_use_date') != today:
                # 新的一天，重置使用次数
                account['last_use_date'] = today
                account['today_use_count'] = 0
            
            # 检查是否超过最大使用次数
            if account.get('today_use_count', 0) < self.max_count:
                return self.current_account_index, account
            
            # 尝试下一个账号
            self.current_account_index = (self.current_account_index + 1) % len(self.accounts)
            attempts += 1
        
        # 所有账号都已达到最大使用次数
        return None, None
    
    def run(self):
        """执行采集任务"""
        for index, item in enumerate(self.collect_items):
            if not self.is_running:
                break
                
            # 检查是否暂停
            while self.is_paused and self.is_running:
                time.sleep(0.1)
                
            if not self.is_running:
                break
            
            # 获取可用账号
            account_index, account = self.get_next_available_account()
            if account is None:
                # 所有账号都已达到最大使用次数
                self.item_collected.emit(index, False, '所有账号均已达到今日最大使用次数', None)
                continue
            
            # 更新账号使用次数
            account['today_use_count'] = account.get('today_use_count', 0) + 1
            self.account_used.emit(account_index, account['today_use_count'])
            
            # 创建API实例
            api = BloggerAPI(account.get('cookies'))
                
            # 更新状态为采集中
            self.status_changed.emit(index, f'采集中-博主信息(账号{account_index+1})')
            
            # 1. 调用API采集博主信息
            user_id = item['user_id']
            success1, message1, data1 = api.get_blogger_info(user_id)
            
            # 采集后切换到下一个账号
            self.current_account_index = (self.current_account_index + 1) % len(self.accounts)
            
            if not success1:
                # 博主信息采集失败，直接发送失败结果
                self.item_collected.emit(index, False, message1, None)
            else:
                # 博主信息采集成功，继续采集数据概览
                self.status_changed.emit(index, '采集中-数据概览')
                
                # 2. 调用API采集数据概览
                success2, message2, data2 = api.get_data_summary(user_id)
                
                # 合并博主信息和数据概览
                combined_data = data1.copy() if data1 else {}
                if success2 and data2:
                    combined_data.update(data2)
                
                # 3. 调用API采集数据表现（根据用户选择的字段）
                self.status_changed.emit(index, '采集中-数据表现')
                success3, message3, data3 = api.get_performance_data(user_id, self.selected_fields)
                
                # 合并数据表现数据
                if success3 and data3:
                    combined_data.update(data3)
                
                # 4. 调用API采集粉丝指标
                self.status_changed.emit(index, '采集中-粉丝指标')
                success4, message4, data4 = api.get_fans_summary(user_id)
                
                # 合并粉丝指标数据
                if success4 and data4:
                    combined_data.update(data4)
                
                # 5. 调用API采集粉丝画像
                self.status_changed.emit(index, '采集中-粉丝画像')
                success5, message5, data5 = api.get_fans_profile(user_id)
                
                # 合并粉丝画像数据
                if success5 and data5:
                    combined_data.update(data5)
                
                # 构建最终消息
                final_message = message1
                if not success2:
                    final_message += f"（数据概览失败: {message2}）"
                if not success3:
                    final_message += f"（数据表现失败: {message3}）"
                if not success4:
                    final_message += f"（粉丝指标失败: {message4}）"
                if not success5:
                    final_message += f"（粉丝画像失败: {message5}）"
                    
                self.item_collected.emit(index, True, final_message, combined_data)
            
            # 等待间隔时间（除了最后一个）
            if index < len(self.collect_items) - 1 and self.is_running:
                time.sleep(1.0)
                
        # 发送完成信号
        if self.is_running:
            self.all_finished.emit()
            
    def pause(self):
        """暂停采集"""
        self.is_paused = True
        
    def resume(self):
        """恢复采集"""
        self.is_paused = False
        
    def stop(self):
        """停止采集"""
        self.is_running = False
        self.is_paused = False


class TextImportDialog(QDialog):
    """文本导入对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.result_text = None
        self.init_ui()
        
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle('文本导入')
        self.resize(600, 400)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 提示标签
        tip = QLabel('请输入URL，每行一个：')
        tip.setStyleSheet('font-size: 14px; color: #666;')
        layout.addWidget(tip)
        
        # 文本输入框
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText(
            '请输入URL，每行一个。\n'
            '支持格式：\n'
            'https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/xxx\n'
            'https://www.xiaohongshu.com/user/profile/xxx'
        )
        layout.addWidget(self.text_edit)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_btn = PushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        
        import_btn = PrimaryPushButton('导入')
        import_btn.clicked.connect(self.accept_import)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(import_btn)
        
        layout.addLayout(button_layout)
        
    def accept_import(self):
        """确认导入"""
        self.result_text = self.text_edit.toPlainText()
        self.accept()


class CollectManagePage(QWidget):
    """采集管理页面"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName('collectManagePage')
        self.collect_items = []  # 存储采集项目
        self.collect_thread = None  # 采集线程
        self.is_collecting = False  # 是否正在采集
        self.settings_file = Path('data/collect_settings.json')
        self.init_ui()
        
    def init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # 顶部按钮区域
        top_layout = QHBoxLayout()
        
        # 导入按钮组
        self.excel_import_btn = PushButton('Excel导入', self, FIF.DOCUMENT)
        self.excel_import_btn.clicked.connect(self.import_from_excel)
        
        self.text_import_btn = PushButton('文本导入', self, FIF.EDIT)
        self.text_import_btn.clicked.connect(self.import_from_text)
        
        self.txt_import_btn = PushButton('TXT文件导入', self, FIF.DOCUMENT)
        self.txt_import_btn.clicked.connect(self.import_from_txt)
        
        # 采集控制按钮组
        self.start_btn = PrimaryPushButton('开始采集', self, FIF.PLAY)
        self.start_btn.clicked.connect(self.start_collect)
        
        self.pause_btn = PushButton('暂停采集', self, FIF.PAUSE)
        self.pause_btn.clicked.connect(self.pause_collect)
        
        self.stop_btn = PushButton('结束采集', self, FIF.CLOSE)
        self.stop_btn.clicked.connect(self.stop_collect)
        
        top_layout.addWidget(self.excel_import_btn)
        top_layout.addWidget(self.text_import_btn)
        top_layout.addWidget(self.txt_import_btn)
        top_layout.addSpacing(20)
        top_layout.addWidget(self.start_btn)
        top_layout.addWidget(self.pause_btn)
        top_layout.addWidget(self.stop_btn)
        top_layout.addStretch()
        
        layout.addLayout(top_layout)
        
        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            '蒲公英主页',
            '小红书主页',
            '小红书ID',
            '博主昵称',
            '采集状态',
            '采集时间'
        ])
        
        # 设置表格样式
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # 去除表格边框黑线
        self.table.setShowGrid(False)
        self.table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: transparent;
            }
            QTableWidget::item {
                border: none;
                padding: 5px;
                color: #000000;
            }
            QTableWidget::item:selected {
                background-color: #e5f3ff;
                color: #000000;
            }
        """)
        
        layout.addWidget(self.table)
        
        # 设置按钮初始状态
        self.update_button_states(is_collecting=False)
        
    def update_button_states(self, is_collecting):
        """更新按钮状态"""
        self.start_btn.setEnabled(not is_collecting)
        self.pause_btn.setEnabled(is_collecting)
        self.stop_btn.setEnabled(is_collecting)
        
    def extract_user_id(self, url):
        """从LURL中提取用户ID"""
        # 匹配蒲公英URL
        pgy_pattern = r'pgy\.xiaohongshu\.com/solar/pre-trade/blogger-detail/([a-f0-9]+)'
        match = re.search(pgy_pattern, url)
        if match:
            return match.group(1)
        
        # 匹配小红书URL
        xhs_pattern = r'www\.xiaohongshu\.com/user/profile/([a-f0-9]+)'
        match = re.search(xhs_pattern, url)
        if match:
            return match.group(1)
        
        return None
        
    def is_valid_url(self, url):
        """验证URL是否有效"""
        return ('pgy.xiaohongshu.com/solar/pre-trade/blogger-detail' in url or
                'www.xiaohongshu.com/user/profile' in url)
                
    def generate_urls(self, user_id):
        """根据ID生成两种URL"""
        pgy_url = f'https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/{user_id}'
        xhs_url = f'https://www.xiaohongshu.com/user/profile/{user_id}'
        return pgy_url, xhs_url
        
    def add_collect_item(self, url):
        """添加采集项目"""
        if not self.is_valid_url(url):
            return False
            
        user_id = self.extract_user_id(url)
        if not user_id:
            return False
            
        # 检查是否已存在
        for item in self.collect_items:
            if item['user_id'] == user_id:
                return False  # 已存在，跳过
                
        pgy_url, xhs_url = self.generate_urls(user_id)
        
        item = {
            'pgy_url': pgy_url,
            'xhs_url': xhs_url,
            'user_id': user_id,
            'nickname': '',  # 博主昵称，采集时填充
            'status': '待采集',
            'collect_time': ''
        }
        
        self.collect_items.append(item)
        return True
        
    def refresh_table(self):
        """刷新表格显示"""
        self.table.setRowCount(len(self.collect_items))
        for i, item in enumerate(self.collect_items):
            self.table.setItem(i, 0, QTableWidgetItem(item['pgy_url']))
            self.table.setItem(i, 1, QTableWidgetItem(item['xhs_url']))
            self.table.setItem(i, 2, QTableWidgetItem(item['user_id']))
            self.table.setItem(i, 3, QTableWidgetItem(item['nickname']))
            self.table.setItem(i, 4, QTableWidgetItem(item['status']))
            self.table.setItem(i, 5, QTableWidgetItem(item['collect_time']))
            
    def import_from_excel(self):
        """从Excel导入"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            '选择Excel文件',
            '',
            'Excel Files (*.xlsx *.xls)'
        )
        
        if not file_path:
            return
            
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            added_count = 0
            skipped_count = 0
            
            # 从第一行开始读取第一列
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
                cell_value = row[0].value
                if cell_value:
                    url = str(cell_value).strip()
                    if self.add_collect_item(url):
                        added_count += 1
                    else:
                        skipped_count += 1
                        
            self.refresh_table()
            
            InfoBar.success(
                title='导入成功',
                content=f'成功导入 {added_count} 条，跳过 {skipped_count} 条',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
        except ImportError:
            InfoBar.error(
                title='错误',
                content='请先安装 openpyxl: pip install openpyxl',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title='导入失败',
                content=f'无法读取Excel文件: {str(e)}',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
    def import_from_text(self):
        """从文本导入"""
        dialog = TextImportDialog(self)
        if dialog.exec_():
            text = dialog.result_text
            if text:
                lines = text.strip().split('\n')
                
                added_count = 0
                skipped_count = 0
                
                for line in lines:
                    url = line.strip()
                    if url:
                        if self.add_collect_item(url):
                            added_count += 1
                        else:
                            skipped_count += 1
                            
                self.refresh_table()
                
                InfoBar.success(
                    title='导入成功',
                    content=f'成功导入 {added_count} 条，跳过 {skipped_count} 条',
                    orient=Qt.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                    parent=self
                )
                
    def import_from_txt(self):
        """从TXT文件导入"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            '选择TXT文件',
            '',
            'Text Files (*.txt)'
        )
        
        if not file_path:
            return
            
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                
            added_count = 0
            skipped_count = 0
            
            for line in lines:
                url = line.strip()
                if url:
                    if self.add_collect_item(url):
                        added_count += 1
                    else:
                        skipped_count += 1
                        
            self.refresh_table()
            
            InfoBar.success(
                title='导入成功',
                content=f'成功导入 {added_count} 条，跳过 {skipped_count} 条',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
        except Exception as e:
            InfoBar.error(
                title='导入失败',
                content=f'无法读取TXT文件: {str(e)}',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
    def start_collect(self):
        """开始采集"""
        if not self.collect_items:
            InfoBar.warning(
                title='提示',
                content='请先导入采集目标',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
            return
            
        if self.is_collecting:
            InfoBar.warning(
                title='提示',
                content='正在采集中，请勿重复操作',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
            return
            
        # 重新加载最新的采集设置（每次采集都重新读取文件）
        print('=' * 50)
        print('开始采集 - 重新加载配置文件...')
        settings = self.load_settings()
        max_count = settings.get('max_count', 9999)
        
        # 获取用户选择的数据表现字段
        selected_fields = settings.get('performance_fields', [])
        print(f'账号最大使用次数: {max_count}')
        print(f'选择的数据表现字段数量: {len(selected_fields)}')
        print(f'选择的字段: {selected_fields}')
        print('=' * 50)
        
        # 验证是否选择了数据表现字段
        if not selected_fields:
            InfoBar.warning(
                title='提示',
                content='未选择任何数据表现字段，请先在采集设置中选择需要采集的字段',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            return
        
        # 从账号管理页获取所有有效账号
        accounts = self.get_valid_accounts()
        if not accounts:
            InfoBar.error(
                title='错误',
                content='没有可用的账号，请先在账号管理中添加并验证账号',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            return
        
        # 保存账号列表的引用，用于后续保存使用记录
        self.current_accounts = accounts
            
        # 创建采集线程
        self.collect_thread = CollectThread(self.collect_items, accounts, max_count, selected_fields)
        self.collect_thread.item_collected.connect(self.on_item_collected)
        self.collect_thread.status_changed.connect(self.on_status_changed)
        self.collect_thread.all_finished.connect(self.on_all_finished)
        self.collect_thread.account_used.connect(self.on_account_used)
        
        self.is_collecting = True
        self.update_button_states(is_collecting=True)
        self.collect_thread.start()
        
        InfoBar.info(
            title='开始采集',
            content=f'开始采集 {len(self.collect_items)} 个目标（已选择 {len(selected_fields)} 种数据表现字段）',
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self
        )
        
    def pause_collect(self):
        """暂停采集"""
        if not self.is_collecting or not self.collect_thread:
            InfoBar.warning(
                title='提示',
                content='当前没有正在进行的采集任务',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
            return
            
        if self.collect_thread.is_paused:
            # 恢复采集
            self.collect_thread.resume()
            self.pause_btn.setText('暂停采集')
            InfoBar.info(
                title='恢复采集',
                content='已恢复采集任务',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
        else:
            # 暂停采集
            self.collect_thread.pause()
            self.pause_btn.setText('恢复采集')
            InfoBar.info(
                title='已暂停',
                content='采集任务已暂停',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
        
    def stop_collect(self):
        """结束采集"""
        if not self.is_collecting or not self.collect_thread:
            InfoBar.warning(
                title='提示',
                content='当前没有正在进行的采集任务',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
            return
            
        # 停止线程
        self.collect_thread.stop()
        self.collect_thread.wait()
        self.is_collecting = False
        self.pause_btn.setText('暂停采集')
        self.update_button_states(is_collecting=False)
        
        InfoBar.success(
            title='已停止',
            content='采集任务已终止',
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=self
        )
        
    def load_settings(self):
        """加载采集设置（每次都从文件重新读取）"""
        default_settings = {
            'save_mode': 'local',
            'local': {
                'filename': 'collected_data.xlsx',
                'path': str(Path.home() / 'Documents')
            },
            'performance_fields': [
                '日常笔记-图文+视频-近30天-全流量',
                '日常笔记-图文-近30天-全流量',
                '日常笔记-视频-近30天-全流量',
                '日常笔记-图文+视频-近90天-全流量',
                '日常笔记-图文-近90天-全流量',
                '日常笔记-视频-近90天-全流量',
                '合作笔记-图文+视频-近30天-全流量',
                '合作笔记-图文-近30天-全流量',
                '合作笔记-视频-近30天-全流量',
                '合作笔记-图文+视频-近90天-全流量',
                '合作笔记-图文-近90天-全流量',
                '合作笔记-视频-近90天-全流量',
            ],
            'max_count': 9999
        }
        
        try:
            if self.settings_file.exists():
                print(f'从文件加载设置: {self.settings_file}')
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    default_settings.update(loaded)
                    print(f'成功加载配置，performance_fields数量: {len(loaded.get("performance_fields", []))}')
            else:
                print(f'配置文件不存在，使用默认设置: {self.settings_file}')
        except Exception as e:
            print(f'加载设置失败: {e}')
            
        return default_settings
        
    def get_valid_accounts(self):
        """获取所有有效的账号"""
        account_file = Path('data/pgy_username.json')
        try:
            if account_file.exists():
                with open(account_file, 'r', encoding='utf-8') as f:
                    accounts = json.load(f)
                    # 返回所有状态为正常的账号
                    valid_accounts = [account for account in accounts if account.get('status') == '正常']
                    return valid_accounts
        except Exception as e:
            print(f'读取账号失败: {e}')
            
        return []
    
    def save_accounts(self, accounts):
        """保存账号数据"""
        account_file = Path('data/pgy_username.json')
        try:
            account_file.parent.mkdir(parents=True, exist_ok=True)
            with open(account_file, 'w', encoding='utf-8') as f:
                json.dump(accounts, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f'保存账号失败: {e}')
        
    def on_account_used(self, account_index, used_count):
        """账号使用回调"""
        # 更新当前账号列表中的使用次数（已在线程中更新）
        # 这里可以添加日志或UI更新
        print(f'账号 {account_index + 1} 今日已使用 {used_count} 次')
    
    def on_status_changed(self, index, status):
        """状态更新"""
        if 0 <= index < len(self.collect_items):
            self.collect_items[index]['status'] = status
            self.refresh_table()
            
    def on_item_collected(self, index, success, message, data):
        """单个采集完成"""
        if 0 <= index < len(self.collect_items):
            item = self.collect_items[index]
            
            if success and data:
                # 更新博主昵称
                item['nickname'] = data.get('name', '')
                item['status'] = '已完成'
                item['collect_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # 保存采集的数据
                item['collected_data'] = data
            else:
                item['status'] = f'失败: {message}'
                item['collect_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
            self.refresh_table()
            
    def on_all_finished(self):
        """所有采集完成"""
        self.is_collecting = False
        self.pause_btn.setText('暂停采集')
        self.update_button_states(is_collecting=False)
        
        # 保存账号使用记录
        if hasattr(self, 'current_accounts'):
            # 读取所有账号
            account_file = Path('data/pgy_username.json')
            try:
                with open(account_file, 'r', encoding='utf-8') as f:
                    all_accounts = json.load(f)
                
                # 更新使用记录
                for current_acc in self.current_accounts:
                    for acc in all_accounts:
                        if acc.get('cookies') == current_acc.get('cookies'):
                            acc['last_use_date'] = current_acc.get('last_use_date', '')
                            acc['today_use_count'] = current_acc.get('today_use_count', 0)
                            break
                
                # 保存回文件
                self.save_accounts(all_accounts)
            except Exception as e:
                print(f'保存账号使用记录失败: {e}')
        
        # 统计结果
        success_count = sum(1 for item in self.collect_items if item['status'] == '已完成')
        failed_count = sum(1 for item in self.collect_items if '失败' in item['status'])
        
        InfoBar.success(
            title='采集完成',
            content=f'成功: {success_count} 个 | 失败: {failed_count} 个',
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self
        )
        
        # 如果选择了本地保存，自动保存
        settings = self.load_settings()
        if settings.get('save_mode') == 'local':
            self.save_to_excel()
            
    def get_performance_field_headers(self, field_prefix):
        """
        获取数据表现字段的所有子字段表头
        :param field_prefix: 字段前缀，如 '数据表现-日常笔记-图文+视频-近30天-全流量'
        :return: 子字段列表
        """
        return [
            f'{field_prefix}-笔记数',
            f'{field_prefix}-内容类目及占比',
            f'{field_prefix}-曝光中位数',
            f'{field_prefix}-阅读中位数',
            f'{field_prefix}-互动中位数',
            f'{field_prefix}-中位点赞量',
            f'{field_prefix}-中位收藏量',
            f'{field_prefix}-中位评论量',
            f'{field_prefix}-中位分享量',
            f'{field_prefix}-中位关注量',
            f'{field_prefix}-互动率',
            f'{field_prefix}-图文3秒阅读率',
            f'{field_prefix}-千赞笔记比例',
            f'{field_prefix}-百赞笔记比例',
            f'{field_prefix}-预估CPM',
            f'{field_prefix}-预估阅读单价',
            f'{field_prefix}-预估互动单价',
            f'{field_prefix}-阅读量来源-发现页',
            f'{field_prefix}-阅读量来源-搜索页',
            f'{field_prefix}-阅读量来源-关注页',
            f'{field_prefix}-阅读量来源-博主个人页',
            f'{field_prefix}-阅读量来源-附近页',
            f'{field_prefix}-阅读量来源-其他',
            f'{field_prefix}-曝光量来源-发现页',
            f'{field_prefix}-曝光量来源-搜索页',
            f'{field_prefix}-曝光量来源-关注页',
            f'{field_prefix}-曝光量来源-博主个人页',
            f'{field_prefix}-曝光量来源-附近页',
            f'{field_prefix}-曝光量来源-其他',
        ]
    
    def get_performance_field_values(self, data, field_prefix):
        """
        获取数据表现字段的所有子字段值
        :param data: 采集的数据字典
        :param field_prefix: 字段前缀，如 '数据表现-日常笔记-图文+视频-近30天-全流量'
        :return: 值列表
        """
        return [
            data.get(f'{field_prefix}-笔记数', ''),
            data.get(f'{field_prefix}-内容类目及占比', ''),
            data.get(f'{field_prefix}-曝光中位数', ''),
            data.get(f'{field_prefix}-阅读中位数', ''),
            data.get(f'{field_prefix}-互动中位数', ''),
            data.get(f'{field_prefix}-中位点赞量', ''),
            data.get(f'{field_prefix}-中位收藏量', ''),
            data.get(f'{field_prefix}-中位评论量', ''),
            data.get(f'{field_prefix}-中位分享量', ''),
            data.get(f'{field_prefix}-中位关注量', ''),
            data.get(f'{field_prefix}-互动率', ''),
            data.get(f'{field_prefix}-图文3秒阅读率', ''),
            data.get(f'{field_prefix}-千赞笔记比例', ''),
            data.get(f'{field_prefix}-百赞笔记比例', ''),
            data.get(f'{field_prefix}-预估CPM', ''),
            data.get(f'{field_prefix}-预估阅读单价', ''),
            data.get(f'{field_prefix}-预估互动单价', ''),
            data.get(f'{field_prefix}-阅读量来源-发现页', ''),
            data.get(f'{field_prefix}-阅读量来源-搜索页', ''),
            data.get(f'{field_prefix}-阅读量来源-关注页', ''),
            data.get(f'{field_prefix}-阅读量来源-博主个人页', ''),
            data.get(f'{field_prefix}-阅读量来源-附近页', ''),
            data.get(f'{field_prefix}-阅读量来源-其他', ''),
            data.get(f'{field_prefix}-曝光量来源-发现页', ''),
            data.get(f'{field_prefix}-曝光量来源-搜索页', ''),
            data.get(f'{field_prefix}-曝光量来源-关注页', ''),
            data.get(f'{field_prefix}-曝光量来源-博主个人页', ''),
            data.get(f'{field_prefix}-曝光量来源-附近页', ''),
            data.get(f'{field_prefix}-曝光量来源-其他', ''),
        ]
            
    def save_to_excel(self):
        """保存为Excel文件"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # 重新加载最新的采集设置
            print('=' * 50)
            print('保存Excel - 重新加载配置文件...')
            settings = self.load_settings()
            filename = settings['local']['filename']
            filepath = Path(settings['local']['path']) / filename
            selected_fields = settings.get('performance_fields', [])
            print(f'保存文件: {filepath}')
            print(f'包含 {len(selected_fields)} 种数据表现字段')
            print('=' * 50)
            
            # 确保文件名以.xlsx结尾
            if not str(filepath).endswith('.xlsx'):
                filepath = filepath.with_suffix('.xlsx')
                
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '采集数据'
            
            # 基础表头（博主信息 + 数据概览 + 粉丝指标 + 粉丝画像）
            base_headers = [
                '博主主页', '达人 ID', '蒲公英主页', '小红书主页',
                '昵称', '性别', '小红书号', '地理位置',
                '粉丝数量', '获赞与收藏', '合作报价-图文笔记',
                '合作报价-视频笔记', '合作报价-最低报价',
                '签约机构', '内容标签', '合作行业',
                # 数据概览字段
                '发布笔记', '内容类目', '数据更新时间',
                '数据概览-笔记数据-日常笔记-曝光中位数', '数据概览-笔记数据-日常笔记-阅读中位数', '数据概览-笔记数据-日常笔记-互动中位数',
                '数据概览-笔记数据-合作笔记-曝光中位数', '数据概览-合作笔记-阅读中位数', '数据概览-笔记数据-合作笔记-互动中位数',
                '数据概览-笔记数据-预估CPM(图文)', '数据概览-笔记数据-预估CPM(视频)',
                '数据概览-笔记数据-预估阅读单价(图文)', '数据概览-笔记数据-预估阅读单价(视频)',
                '数据概览-笔记数据-预估互动单价(图文)', '数据概览-笔记数据-预估互动单价(视频)',
                '数据概览-笔记数据-预估外溢进店单价(图文)', '数据概览-笔记数据-预估外溢进店单价(视频)',
                '近7天活跃天数', '邀约48小时回复率', '粉丝量变化幅度',
            ]
            
            # 根据用户选择添加数据表现字段（已在前面加载）
            performance_headers = []
            for field in selected_fields:
                field_prefix = f'数据表现-{field}'
                performance_headers.extend(self.get_performance_field_headers(field_prefix))
            
            # 粉丝指标和粉丝画像字段
            fans_headers = [
                # 粉丝指标字段
                '粉丝指标-粉丝增量', '粉丝指标-粉丝量变化幅度', '粉丝指标-活跃粉丝占比', '粉丝指标-阅读粉丝占比', '粉丝指标-互动粉丝占比', '粉丝指标-下单粉丝占比',
                # 粉丝画像字段
                '粉丝画像-性别分布', '粉丝画像-年龄分布', '粉丝画像-地域分布-按省份', '粉丝画像-地域分布-按城市', '粉丝画像-用户设备分布', '粉丝画像-用户兴趣',
                # 采集时间
                '采集时间',
            ]
            
            # 合并所有表头
            headers = base_headers + performance_headers + fans_headers
            ws.append(headers)
            
            # 写入数据
            for item in self.collect_items:
                if item['status'] == '已完成' and 'collected_data' in item:
                    data = item['collected_data']
                    
                    # 基础数据行
                    base_row = [
                        item['pgy_url'],  # 博主主页
                        item['user_id'],  # 达人 ID
                        item['pgy_url'],  # 蒲公英主页
                        item['xhs_url'],  # 小红书主页
                        data.get('name', ''),  # 昵称
                        data.get('gender', ''),  # 性别
                        data.get('redId', ''),  # 小红书号
                        data.get('location', ''),  # 地理位置
                        data.get('fansCount', 0),  # 粉丝数量
                        data.get('likeCollectCountInfo', 0),  # 获赞与收藏
                        data.get('picturePrice', 0.0),  # 合作报价-图文笔记
                        data.get('videoPrice', 0.0),  # 合作报价-视频笔记
                        data.get('lowerPrice', 0.0),  # 合作报价-最低报价
                        data.get('noteSign', ''),  # 签约机构
                        data.get('contentTags', ''),  # 内容标签
                        data.get('tradeType', ''),  # 合作行业
                        # 数据概览字段
                        data.get('noteNumber', ''),  # 发布笔记
                        data.get('noteType', ''),  # 内容类目
                        data.get('dateKey', ''),  # 数据更新时间
                        data.get('daily_mAccumImpNum', ''),  # 日常笔记-曝光中位数
                        data.get('daily_mValidRawReadFeedNum', ''),  # 日常笔记-阅读中位数
                        data.get('daily_mEngagementNum', ''),  # 日常笔记-互动中位数
                        data.get('coop_mAccumImpNum', ''),  # 合作笔记-曝光中位数
                        data.get('coop_mValidRawReadFeedNum', ''),  # 合作笔记-阅读中位数
                        data.get('coop_mEngagementNum', ''),  # 合作笔记-互动中位数
                        data.get('estimatePictureCpm', ''),  # 预估CPM(图文)
                        data.get('estimateVideoCpm', ''),  # 预估CPM(视频)
                        data.get('picReadCost', ''),  # 预估阅读单价(图文)
                        data.get('videoReadCostV2', ''),  # 预估阅读单价(视频)
                        data.get('estimatePictureEngageCost', ''),  # 预估互动单价(图文)
                        data.get('estimateVideoEngageCost', ''),  # 预估互动单价(视频)
                        data.get('estimatePictureCpuv', ''),  # 预估外溢进店单价(图文)
                        data.get('estimateVideoCpuv', ''),  # 预估外溢进店单价(视频)
                        data.get('activeDayInLast7', ''),  # 近7天活跃天数
                        data.get('responseRate', ''),  # 邀约48小时回复率
                        data.get('fans30GrowthBeyondRate', ''),  # 粉丝量变化幅度
                    ]
                    
                    # 根据用户选择添加数据表现字段的值
                    performance_values = []
                    for field in selected_fields:
                        field_prefix = f'数据表现-{field}'
                        performance_values.extend(self.get_performance_field_values(data, field_prefix))
                    
                    # 粉丝指标和粉丝画像数据
                    fans_values = [
                        # 粉丝指标
                        data.get('粉丝指标-粉丝增量', ''),
                        data.get('粉丝指标-粉丝量变化幅度', ''),
                        data.get('粉丝指标-活跃粉丝占比', ''),
                        data.get('粉丝指标-阅读粉丝占比', ''),
                        data.get('粉丝指标-互动粉丝占比', ''),
                        data.get('粉丝指标-下单粉丝占比', ''),
                        # 粉丝画像
                        data.get('粉丝画像-性别分布', ''),
                        data.get('粉丝画像-年龄分布', ''),
                        data.get('粉丝画像-地域分布-按省份', ''),
                        data.get('粉丝画像-地域分布-按城市', ''),
                        data.get('粉丝画像-用户设备分布', ''),
                        data.get('粉丝画像-用户兴趣', ''),
                        # 采集时间
                        item.get('collect_time', ''),
                    ]
                    
                    # 合并所有行数据
                    row = base_row + performance_values + fans_values
                    ws.append(row)
                    
            # 保存文件
            wb.save(filepath)
            
            InfoBar.success(
                title='保存成功',
                content=f'数据已保存到: {filepath}',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
        except ImportError:
            InfoBar.error(
                title='错误',
                content='请先安装 openpyxl: pip install openpyxl',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title='保存失败',
                content=f'无法保存文件: {str(e)}',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            
        except ImportError:
            InfoBar.error(
                title='错误',
                content='请先安装 openpyxl: pip install openpyxl',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title='保存失败',
                content=f'无法保存文件: {str(e)}',
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
